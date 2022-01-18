from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from roster import Location, Court, Match

_ORDINAL_INDICATORS = ('th', 'st', 'nd', 'rd')
_DATE_FORMAT = '%d/%m/%Y'
_TIME_FORMAT = '%I:%M %p'
_ROW_HEIGHT = 17
_DATE_CELL = 'C4'
_FIRST_TABLE_ROW = 6
_COURT_COLUMN = _TIME_COLUMN = 'B'
_LOCATION_COLUMN = _TEAM_1_COLUMN = 'C'
_TEAM_2_COLUMN = 'D'
_GRADE_COLUMN = 'E'
_REFEREE_COLUMN_1 = 'F'
_REFEREE_COLUMN_2 = 'G'
_EXCEL_COURT_OVERFLOW = 5


class _MatchChanges:
    def __init__(self):
        self.added = {}
        self.removed = {}

    def add(self, match):
        """Registers a match as added

        Args:
            match(Match): The match

        """
        self.added[match.team1] = match

    def remove(self, match):
        """Registers a match as removed

        Args:
            match(Match): The match

        """
        self.removed[match.team1] = match

    def __str__(self):
        removed = self.removed.copy()
        added = self.added.copy()
        moved = []
        for removed_match in self.removed.values():
            # If the match has been removed but not moved, ignore it
            if removed_match.team1 not in self.added:
                continue

            # If the match has already been registed as moved, skip it
            if removed_match.team1 not in removed:
                continue

            # Register a moved match
            moved_ = ([removed.pop(removed_match.team1)], [added.pop(removed_match.team1)])
            if removed_match.team2 not in self.added:
                moved.append(moved_)
                continue

            # Register the moved match as a pair
            moved_[0][0].append(removed.pop(removed_match.team2))
            moved_[0][1].append(added.pop(removed_match.team2))
            moved.append(moved_)

        return f'Moved:\n{moved}\n\nAdded:\n{added}\n\nRemoved:\n{removed}'


def _ordinal(date):
    """Gets the correct ordinal indicator for a date

    Args:
        date(datetime): The date to get the ordinal indicator for

    Returns:
        str: The correct ordinal indicator

    """
    day = int(date.strftime('%d'))

    if day % 10 in (1, 2, 3) and day not in (11, 12, 13):
        return _ORDINAL_INDICATORS[day % 10]

    return _ORDINAL_INDICATORS[0]


def _add_styles(wb):
    """Add styles to a given Excel workbook

    Args:
        wb(Workbook): The Excel workbook

    """
    # Define global style components
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')

    # Define table style components
    table_side = Side(style='thin', color="000000")
    table_border = Border(left=table_side, top=table_side, right=table_side, bottom=table_side)

    # Define time style components
    time_font = Font(bold=True, color='FFFFFF')
    time_fill = PatternFill(bgColor='000000', fill_type='solid')

    # Define forfeit style component
    forfeit_fill = PatternFill(bgColor='FFFF00', fill_type='solid')

    # Define and add named styles
    wb.add_named_style(NamedStyle(name='date', font=bold_font, alignment=center_alignment))
    wb.add_named_style(NamedStyle(name='court', font=bold_font, alignment=center_alignment))
    wb.add_named_style(NamedStyle(name='location', font=bold_font))
    wb.add_named_style(NamedStyle(name='time', font=time_font, alignment=center_alignment, fill=time_fill))
    wb.add_named_style(NamedStyle(name='team', font=bold_font, border=table_border))
    wb.add_named_style(NamedStyle(name='grade', font=bold_font, border=table_border, alignment=center_alignment))
    wb.add_named_style(NamedStyle(name='referee', border=table_border))
    wb.add_named_style(NamedStyle(name='forfeit', fill=forfeit_fill))


def _update_court_header(ws, row, court, location):
    """Update a row in an Excel spreadsheet with a match

    Args:
        ws(Worksheet): The Excel worksheet
        row(int): The row of the Excel spreadsheet to update
        court(Court): The court to update the header with
        location(Location): The location to update the header with

    """
    # Fill in court cell
    court_cell = ws[f'{_COURT_COLUMN}{row}']
    court_cell.value = f'Crt {court.value}'
    court_cell.style = 'court'

    # Fill in location cell
    location_cell = ws[f'{_LOCATION_COLUMN}{row}']
    location_cell.value = str(location)
    location_cell.style = 'location'

    # Increase height of the header row
    ws.row_dimensions[row].height = _ROW_HEIGHT


def _update_row(ws, row, match):
    """Update a row in an Excel spreadsheet with a match

    Args:
        ws(Worksheet): The Excel worksheet
        row(int): The row of the Excel spreadsheet to update
        match(Match): The match to update the row with

    """
    # Fill in match time cell
    match_time_cell = ws[f'{_TIME_COLUMN}{row}']
    match_time_cell.value = match.time.strftime(_TIME_FORMAT).lstrip('0')
    match_time_cell.style = 'time'

    # Fill in match team 1 cell
    match_team_1_cell = ws[f'{_TEAM_1_COLUMN}{row}']
    match_team_1_cell.value = match.team1
    match_team_1_cell.style = 'team'

    # Fill in match team 2 cell
    match_team_2_cell = ws[f'{_TEAM_2_COLUMN}{row}']
    match_team_2_cell.value = match.team2
    match_team_2_cell.style = 'team'

    # Fill in match grade cell
    match_grade_cell = ws[f'{_GRADE_COLUMN}{row}']
    match_grade_cell.value = match.grade
    match_grade_cell.style = 'grade'

    # Format referee cells
    referee_cells = ws[f'{_REFEREE_COLUMN_1}{row}':f'{_REFEREE_COLUMN_2}{row}'][0]
    for referee_cell in referee_cells:
        referee_cell.style = 'referee'

    ws.row_dimensions[row].height = _ROW_HEIGHT


def create_excel(roster, template_location, save_location):
    """Creates an Excel document from a roster

    Args:
        roster(Roster): The roster to export
        template_location(str): The location of the template Excel document
        save_location(str): The location of the folder to save the Excel document to

    """
    wb = load_workbook(template_location)
    _add_styles(wb)
    data = roster.to_dictionary()

    date = data['Date']
    for location in Location:
        ws = wb[str(location)]
        row = _FIRST_TABLE_ROW

        # Fill in the date cell
        date_cell = ws[_DATE_CELL]
        date_cell.value = date.strftime(_DATE_FORMAT).lstrip('0')
        date_cell.style = 'date'

        # Fill in court tables
        for court in Court:
            matches = data[location][court]
            if len(matches) == 0:
                continue

            _update_court_header(ws, row, court, location)
            row += 1

            # Fill in match rows
            for match in matches:
                _update_row(ws, row, match)
                row += 1

    # Set the first worksheet as active
    wb.active = 0

    # Save and close the worksheet
    filename = date.strftime('%d{} %b %Y').format(_ordinal(date)).lstrip('0') + '.xlsx'
    wb.save(f'{save_location}/{filename}')
    wb.close()


def get_excel_date(excel_location):
    """Gets the date for an Excel document

    Args:
        excel_location(str): The location of the Excel document

    """
    wb = load_workbook(excel_location)
    ws = wb.worksheets[0]
    date_string = ws[_DATE_CELL].value
    wb.close()
    return date_string


def _get_row_end(ws, row):
    """Get the column of the end of a row

    Args:
        ws(Worksheet): The Excel worksheet
        row(int): The row number

    Returns:
        str: The column letter of the end of the row

    """
    column = _REFEREE_COLUMN_2
    while ws[f'{chr(ord(column) + 1)}{row}'].border.right.style is not None:
        column = chr(ord(column) + 1)

    return column


def _get_column_end(ws):
    """Get the row at the end of a column

    Args:
        ws(Worksheet): The Excel worksheet

    Returns:
        int: The row number of the end of the column

    """
    row = _FIRST_TABLE_ROW
    while ws[f'{_TIME_COLUMN}{row + 1}'].value is not None:
        row += 1

    return row


def _excel_row_to_match(ws, row, location, court, teams_swapped):
    """Convert a row of an Excel worksheet to a match

    Args:
        ws(Worksheet): The Excel worksheet
        row(int): The row number
        teams_swapped(bool): Whether to swap the teams

    Returns:
        Match: The match

    """
    time = ws[f'{_TIME_COLUMN}{row}']
    team1 = ws[f'{_TEAM_1_COLUMN}{row}']
    team2 = ws[f'{_TEAM_2_COLUMN}{row}']
    grade = ws[f'{_GRADE_COLUMN}{row}']
    return Match(grade, team2 if teams_swapped else team1, team1 if teams_swapped else team2, time, location, court)


def _insert_row(ws, row):
    """Insert a new row in the Excel spreadsheet below a given row

    Args:
        ws(Worksheet): The Excel worksheet
        row(int): The row number

    """
    ws.insert_rows(row)


def _fix_row_heights(ws):
    """Reset the height of the rows in a worksheet

    Args:
        ws(Worksheet): The Excel worksheet

    """
    for i in range(_FIRST_TABLE_ROW, _get_column_end(ws) + 1):
        ws.row_dimensions[i].height = _ROW_HEIGHT


def update_excel(roster, excel_location):
    """Updates an Excel document with a new roster

    Args:
        roster(Roster): The new roster to update the Excel document with
        excel_location(str): The location of the Excel document to update

    """
    wb = load_workbook(excel_location)

    data = roster.to_dictionary()
    match_changes = _MatchChanges()
    for location in Location:
        ws = wb[str(location)]
        data_courts = data[location]
        eof = False

        # Tracking variables for the Excel document
        excel_row = _FIRST_TABLE_ROW
        excel_court = Court.from_string(ws[f'{_COURT_COLUMN}{_FIRST_TABLE_ROW}'].value)

        # Tracking variables for the data object
        data_row = 0
        data_court = Court.COURT_1

        # Iterate through the rows of the worksheet and compare with the data object
        while True:
            # Handle the next court or the end of the worksheet being reached
            court_cell = ws[f'{_COURT_COLUMN}{excel_row}']
            is_court_cell = 'Crt' in court_cell.value if not (court_cell.value is None or court_cell.is_date) else False
            if excel_row != _FIRST_TABLE_ROW and (court_cell.value is None or is_court_cell):
                for i in range(data_row, len(data_courts[data_court])):
                    match = data_courts[data_court][i]
                    _insert_row(ws, excel_row)
                    _update_row(ws, excel_row, match)
                    match_changes.add(match)
                    excel_row += 1

                # Update counters and break if the end of the worksheet has been reached
                data_court = Court.from_num(data_court.value + 1)
                data_row = 0

                court_cell = ws[f'{_COURT_COLUMN}{excel_row}']
                if court_cell.value is not None:
                    excel_court = Court.from_string(court_cell.value)
                else:
                    excel_court = list(Court)[-1]
                    eof = True

            # Handle when the data object has more courts than the Excel document
            while data_court is not None and (data_court.value < excel_court.value or eof):
                if not data_courts[data_court]:
                    data_court = Court.from_num(data_court.value + 1)
                    continue

                # Add header
                _insert_row(ws, excel_row)
                _update_court_header(ws, excel_row, data_court, location)
                excel_row += 1

                # Add matches
                for match in data_courts[data_court]:
                    _insert_row(ws, excel_row)
                    _update_row(ws, excel_row, match)
                    match_changes.add(match)
                    excel_row += 1

                data_court = Court.from_num(data_court.value + 1)
                data_row = 0

            # Exit the loop if the end of the spreadsheet is reached
            if eof:
                break

            # Skip past the court header
            if is_court_cell:
                excel_row += 1

            # Get matches from the data object for comparison
            data_matches = data[location][data_court]

            # Compare the rows in the Excel worksheet with the data object and act accordingly
            data_match = data_matches[data_row]
            data_time = data_match.time

            # Parse the time cell
            excel_time_cell = ws[f'{_TIME_COLUMN}{excel_row}']
            if excel_time_cell.is_date:
                excel_time = excel_time_cell.value
            else:
                excel_time = datetime.strptime(excel_time_cell.value, _TIME_FORMAT)

            # Check if comparing the same time slot
            if data_time == excel_time:
                # Values for each relevant cell in the current excel row
                excel_team1 = ws[f'{_TEAM_1_COLUMN}{excel_row}']
                excel_team2 = ws[f'{_TEAM_2_COLUMN}{excel_row}']
                excel_grade = ws[f'{_GRADE_COLUMN}{excel_row}']
                excel_team1_text = str(excel_team1.value)
                excel_team2_text = str(excel_team2.value)
                excel_grade_text = str(excel_grade.value)

                # Compare team 1
                if data_match.team1 != excel_team1_text:
                    removed_match = _excel_row_to_match(ws, excel_row, location, excel_court, False)
                    match_changes.remove(removed_match)
                    match_changes.add(data_match)
                    excel_team1.value = data_match.team1

                # Compare team 2
                if data_match.team2 != excel_team2_text:
                    removed_match = _excel_row_to_match(ws, excel_row, location, excel_court, True)
                    match_changes.remove(removed_match)
                    match_changes.add(data_match)
                    excel_team2.value = data_match.team2

                # Compare grade
                if data_match.grade != excel_grade_text:
                    excel_grade.value = data_match.grade

                data_row += 1
                excel_row += 1
            else:
                if excel_time in [match.time for match in data_matches]:
                    while data_matches[data_row].time != excel_time:
                        new_data_match = data_matches[data_row]
                        _insert_row(ws, excel_row)
                        _update_row(ws, excel_row, new_data_match)
                        match_changes.add(new_data_match)
                        data_row += 1
                        excel_row += 1
                else:
                    match = _excel_row_to_match(ws, excel_row, location, excel_court, False)
                    row_cells = ws[f'{_TEAM_1_COLUMN}{excel_row}':f'{_get_row_end(ws, excel_row)}{excel_row}'][0]

                    # Highlight the cells yellow
                    for row_cell in row_cells:
                        row_cell.fill = PatternFill(fgColor='FFFF00', fill_type='solid')

                    # Replace the team names with a forfeit label
                    row_cells[0].value = row_cells[1].value = 'FORFEIT'

                    match_changes.remove(match)
                    excel_row += 1

        _fix_row_heights(ws)

    # Set the first worksheet as active
    wb.active = 0
    
    wb.save(excel_location)
    wb.close()
